import string

from openpyxl.styles import PatternFill, Font, Side, Border, NamedStyle, Alignment
from openpyxl.workbook import Workbook


def create_excel(file_path: str):
    wb = Workbook()
    del wb['Sheet']  # 删除初始Sheet表
    sheet = wb.create_sheet(title='新建')
    fill = PatternFill("solid", fgColor="92CDDC")  # 设置填充色
    font = Font(b=True)  # 字体
    side = Side(style='thin', color='000000')  # 边框样式
    alignment = Alignment(horizontal='center', vertical='center')  # 文字居中
    border = Border(left=side, right=side, top=side, bottom=side)  # 设置边框
    sty_title = NamedStyle(
        name='title', font=font, fill=fill, border=border, alignment=alignment, number_format='General')  # 格式
    titles = ['编号', '姓名', '身份证号', '邮箱']
    for index, value in enumerate(titles):
        sheet.cell(column=index + 1, row=1, value=value).style = sty_title
    sheet.freeze_panes = 'A2'  # 首行固定
    # 设置列宽
    for s in string.ascii_uppercase[:4]:
        sheet.column_dimensions[s].width = 20
    # 设置行高
    for n in range(1, sheet.max_row + 1):
        sheet.row_dimensions[n].height = 20

    data = [
        {'number': 1, 'name': '李四', 'id_card': '410222199504120000', 'mail': 'lisi@139.com'},
        {'number': 2, 'name': '张三', 'id_card': '410222199504121111', 'mail': 'zhangsan@139.com'}
    ]
    i = 2
    for info in data:
        sheet.cell(row=i, column=1, value=info.get('number'))
        sheet.cell(row=i, column=2, value=info.get('name'))
        sheet.cell(row=i, column=3, value=info.get('id_card'))
        sheet.cell(row=i, column=4, value=info.get('mail'))
        i += 1
    wb.save(file_path)
    return


def update_excel(file_path: str):
    """更新excel文件"""
    pass
